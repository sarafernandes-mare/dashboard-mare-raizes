"""
Processa histórico Hotmart do Maré Educação e gera XLSX com:
- Aba 1: Parcelas por Venda (X = paga, vermelho = atrasada, cinza = futura)
- Aba 2: Projeção 12 meses (líquido)
- Aba 3: Inadimplência (métricas)
"""
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

import glob as _glob
import os as _os

PASTA = r"C:\Users\saram\Downloads"
SRC = r"C:\Users\saram\Downloads\historico mare 11 meses 2026.csv"  # base histórica fixa

# Auto-detecta TODOS os exports da Hotmart na pasta: arquivos .xlsx ou .xls
# cujo nome contém "venda" OU "sales_history" (case-insensitive).
# Combina todos com o histórico e remove duplicatas pelo código da transação.
def _eh_export(p):
    n = _os.path.basename(p).lower()
    if n.startswith("~$"):       # arquivo de lock do Excel (file aberto)
        return False
    if "base_mare" in n:          # não pegar o próprio output
        return False
    return ("venda" in n) or ("sales_history" in n) or ("sales-history" in n)

# TODOS os exports válidos da pasta E SUBPASTAS (recursivo).
# A validação por aba 'Report' filtra os errados rapidamente.
SRC_NOVOS = sorted([f for f in (_glob.glob(_os.path.join(PASTA, "**", "*.xlsx"), recursive=True)
                                + _glob.glob(_os.path.join(PASTA, "**", "*.xls"), recursive=True))
                    if _eh_export(f)],
                   key=_os.path.getmtime)  # do mais antigo p/ o mais novo
if SRC_NOVOS:
    print(f"Exports Hotmart detectados: {len(SRC_NOVOS)}")
    for _f in SRC_NOVOS:
        print(f"  - {_os.path.basename(_f)}")
else:
    print("Nenhum export Hotmart encontrado (arquivo xlsx com 'venda' no nome).")

OUT = r"C:\Users\saram\Downloads\base_mare_parcelas_projecao.xlsx"
HOJE = datetime(2026, 5, 18)  # provisório; recalculado p/ a transação mais recente após ler os dados
PRODUTOS_RAIZES = ['Raízes', 'Raízes - R']

# Fills
FILL_PAGA = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ATRASADA = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_FUTURA = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_FORA = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_DESC = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def parse_data(serie):
    """Aceita dd/mm/aaaa HH:MM e dd/mm/aaaa HH:MM:SS."""
    d = pd.to_datetime(serie, format='%d/%m/%Y %H:%M:%S', errors='coerce')
    falta = d.isna()
    if falta.any():
        d2 = pd.to_datetime(serie, format='%d/%m/%Y %H:%M', errors='coerce')
        d = d.fillna(d2)
    return d

print("Lendo histórico antigo (CSV)...")
df_old = pd.read_csv(SRC, sep=';', encoding='utf-8', low_memory=False)

dfs = [df_old]
for _f in SRC_NOVOS:
    try:
        _dfn = pd.read_excel(_f, sheet_name='Report')
    except Exception as _e:
        print(f"  [PULOU] {_os.path.basename(_f)} (não tem aba 'Report')")
        continue
    # Valida: precisa ter a estrutura do relatório Hotmart (≥ 60 colunas)
    if len(_dfn.columns) < 60:
        print(f"  [PULOU] {_os.path.basename(_f)} (não parece export Hotmart — só {len(_dfn.columns)} colunas)")
        continue
    # Alinha colunas: usa só as comuns ao histórico (corta extras)
    _ncols = min(len(df_old.columns), len(_dfn.columns))
    _dfn = _dfn.iloc[:, :_ncols].copy()
    _dfn.columns = df_old.columns[:_ncols]
    if 'Código da transação' not in _dfn.columns:
        print(f"  [PULOU] {_os.path.basename(_f)} (sem coluna 'Código da transação')")
        continue
    print(f"  Lendo export: {_os.path.basename(_f)} ({len(_dfn)} linhas, {len(_dfn.columns)} colunas)")
    dfs.append(_dfn)

# Combinar todos e remover duplicatas pelo Código da transação (mantém a 1ª = a mais antiga)
df = pd.concat(dfs, ignore_index=True)
antes = len(df)
df = df.drop_duplicates(subset=['Código da transação'], keep='first').reset_index(drop=True)
print(f"  Combinado: {antes} linhas -> {len(df)} após remover {antes - len(df)} duplicatas")

df['Data da transação'] = parse_data(df['Data da transação'])
df['Faturamento líquido'] = pd.to_numeric(df['Faturamento líquido'], errors='coerce')
df['Quantidade de cobranças'] = pd.to_numeric(df['Quantidade de cobranças'], errors='coerce')
df['Quantidade total de parcelas'] = pd.to_numeric(df['Quantidade total de parcelas'], errors='coerce')

# Normalizar produto
df['Produto'] = df['Produto'].astype(str).str.strip()
df['Valor de compra com impostos'] = pd.to_numeric(df['Valor de compra com impostos'], errors='coerce')

# Remover transações que NÃO são vendas reais (tentativas falhas, cancelamentos, etc.)
_lixo = ['Cancelado', 'Expirado', 'Boleto impresso', 'Reembolsado',
         'Reclamado', 'Iniciada', 'Chargeback']
_antes = len(df)
df = df[~df['Status da transação'].isin(_lixo)].reset_index(drop=True)
print(f"  Limpeza de status não-vendas: {_antes} -> {len(df)} (removidos {_antes - len(df)} cancelados/expirados/etc.)")

# Data de referência = transação mais recente dos dados (robusto p/ automação)
_maxd = df['Data da transação'].max()
if pd.notna(_maxd):
    HOJE = datetime(_maxd.year, _maxd.month, _maxd.day)
print(f"  Data de referência (transação mais recente): {HOJE.strftime('%d/%m/%Y')}")

# Filtrar só produtos Raízes
n_antes = len(df)
df = df[df['Produto'].isin(PRODUTOS_RAIZES)].reset_index(drop=True)
print(f"  Filtro Raízes: {n_antes} -> {len(df)} linhas (removidos {n_antes - len(df)} de outros produtos)")

# ID da venda
df['ID_venda'] = df['Código do assinante']
mask = df['ID_venda'].isna() | (df['ID_venda'].astype(str).isin(['(none)', 'nan']))
df.loc[mask, 'ID_venda'] = df.loc[mask, 'Código da transação']

PRECO_RAIZES_AVISTA = 2997.0  # aprox. preço à vista para inferir N de Assinatura
DIAS_INATIVIDADE_CANCEL = 45  # após N dias sem atividade, plano é considerado encerrado

def total_parcelas_teto(tipo, qtd_total):
    """Retorna o teto máximo teórico de parcelas (antes de ajuste dinâmico)."""
    if pd.isna(tipo):
        return 1
    t = str(tipo)
    if t == 'Assinatura':
        return 12  # teto — ajuste dinâmico depois
    if t.startswith('Parcelamento padr'):
        return 1
    if t.startswith('Apenas'):
        return 1
    if t.startswith('Parcelado SEM'):
        return 1
    if t == 'Parcelado Hotmart - Checkout':
        return 1
    if pd.isna(qtd_total) or qtd_total == 0:
        return 1
    return int(qtd_total)


def total_parcelas_efetivo(tipo, qtd_total, valor_bruto, max_cob_obs, ultima_atividade, hoje):
    """Estima o total real de parcelas do plano:
    - Assinatura: infere N a partir do valor da parcela vs preço do produto;
      se inativa há muito tempo, usa max_cob observado.
    - Recuperador/Parcelado: usa Quantidade total de parcelas, mas se inativo corta em max_cob.
    - À vista: 1.
    """
    teto = total_parcelas_teto(tipo, qtd_total)
    if teto == 1:
        return 1

    dias_inativo = (hoje - ultima_atividade).days if ultima_atividade else 999

    if str(tipo) == 'Assinatura':
        # Infere N pelo valor bruto da parcela (R$ preço / R$ parcela)
        if valor_bruto and valor_bruto > 0:
            n_inferido = round(PRECO_RAIZES_AVISTA / valor_bruto)
            n_inferido = max(1, min(12, n_inferido))
        else:
            n_inferido = max_cob_obs
        # Se plano inativo há > 45 dias, considera que terminou em max_cob
        if dias_inativo > DIAS_INATIVIDADE_CANCEL:
            return max(max_cob_obs, n_inferido) if max_cob_obs >= n_inferido else max_cob_obs
        # Plano ativo: usa N inferido (respeitando max_cob já observado)
        return max(n_inferido, max_cob_obs)

    # Recuperador / Parcelado Hotmart
    if dias_inativo > DIAS_INATIVIDADE_CANCEL and max_cob_obs < teto:
        return max_cob_obs
    return teto

print("Processando vendas...")
vendas = []
for id_venda, g in df.groupby('ID_venda'):
    g = g.sort_values('Data da transação')
    first = g.iloc[0]
    tipo = first['Tipo de cobrança']

    # Descobrir min_cob, max_cob e calcular data da parcela 1 retroativamente
    cobs_validas = g[g['Quantidade de cobranças'].notna()]
    if len(cobs_validas) == 0:
        min_cob = 1
        max_cob_obs = 0
        data_venda = g['Data da transação'].min()
    else:
        min_cob = int(cobs_validas['Quantidade de cobranças'].min())
        max_cob_obs = int(cobs_validas['Quantidade de cobranças'].max())
        data_min_cob = cobs_validas[cobs_validas['Quantidade de cobranças'] == min_cob]['Data da transação'].min()
        data_venda = data_min_cob - relativedelta(months=min_cob - 1)

    ultima_atividade = g['Data da transação'].max()
    v_bruto = first.get('Valor de compra com impostos')
    v_bruto = float(v_bruto) if v_bruto is not None and not pd.isna(v_bruto) else 0.0
    total = total_parcelas_efetivo(tipo, first['Quantidade total de parcelas'],
                                   v_bruto, max_cob_obs, ultima_atividade, HOJE)

    # Valor líquido por parcela: pega de uma linha paga; se não tiver, da primeira
    pagas_any = g[g['Status da transação'].isin(['Completo', 'Aprovado'])]
    if len(pagas_any) > 0 and not pd.isna(pagas_any.iloc[0]['Faturamento líquido']):
        valor_liq = float(pagas_any.iloc[0]['Faturamento líquido'])
    else:
        v = first['Faturamento líquido']
        valor_liq = float(v) if not pd.isna(v) else 0.0

    # Status de cada parcela
    parcelas = {}
    for n in range(1, total + 1):
        data_esperada = data_venda + relativedelta(months=n - 1)
        if n < min_cob:
            # Parcela anterior ao período observado do CSV
            parcelas[n] = ('fora_periodo', data_esperada, valor_liq, None)
            continue
        linhas_n = g[g['Quantidade de cobranças'] == n]
        pagou = linhas_n[linhas_n['Status da transação'].isin(['Completo', 'Aprovado'])]
        # Atrasado = falhou; Aguardando Pagto = boleto/pix emitido e não pago (também não recebido)
        atrasado_tent = linhas_n[linhas_n['Status da transação'].isin(['Atrasado', 'Aguardando Pagto'])]

        if len(pagou) > 0:
            data_pgto = pagou['Data da transação'].min()
            parcelas[n] = ('paga', data_esperada, valor_liq, data_pgto)
        elif len(atrasado_tent) > 0:
            # Hotmart tentou cobrar e não recebeu (atrasado ou aguardando pagamento)
            parcelas[n] = ('atrasada', data_esperada, valor_liq, None)
        elif data_esperada > HOJE:
            parcelas[n] = ('futura', data_esperada, valor_liq, None)
        else:
            # Data passou mas NENHUM registro no CSV (cancelamento? pago fora?)
            parcelas[n] = ('desconhecida', data_esperada, valor_liq, None)

    metodo = first.get('Método de pagamento')
    metodo = str(metodo) if metodo is not None and not pd.isna(metodo) else 'Não informado'

    vendas.append({
        'id_venda': str(id_venda),
        'email': first['Email do(a) Comprador(a)'],
        'nome': first['Comprador(a)'],
        'produto': first['Produto'],
        'tipo': tipo,
        'metodo': metodo,
        'data_venda': data_venda,
        'total_parcelas': total,
        'valor_liquido_parcela': valor_liq,
        'parcelas': parcelas,
    })

vendas.sort(key=lambda v: (v['email'] or '', v['data_venda'] or datetime.min))
print(f"Total de vendas identificadas: {len(vendas)}")

# ========== Métricas agregadas (reusadas em várias abas) ==========
total_vendas = len(vendas)
vendas_com_atraso = sum(1 for v in vendas if any(s == 'atrasada' for s, _, _, _ in v['parcelas'].values()))
emails_unicos = set(v['email'] for v in vendas)
emails_inad = set(v['email'] for v in vendas if any(s == 'atrasada' for s, _, _, _ in v['parcelas'].values()))

total_parc_pagas = 0
total_parc_atrasadas = 0
total_parc_desc = 0
valor_em_atraso = 0.0
valor_recebido = 0.0
valor_desc = 0.0
for v in vendas:
    for n, (status, data_esp, valor, data_pgto) in v['parcelas'].items():
        if status == 'paga':
            total_parc_pagas += 1
            valor_recebido += valor
        elif status == 'atrasada':
            total_parc_atrasadas += 1
            valor_em_atraso += valor
        elif status == 'desconhecida':
            total_parc_desc += 1
            valor_desc += valor
total_parc_venc = total_parc_pagas + total_parc_atrasadas

# ========== Taxa de inadimplência histórica por forma de pagamento ==========
# Base: parcelas já vencidas (pagas + atrasadas confirmadas). Não conta futuras/desconhecidas/fora.
from collections import defaultdict
hist = defaultdict(lambda: {'pagas': 0, 'atrasadas': 0})       # por Método de pagamento
hist_tipo = defaultdict(lambda: {'pagas': 0, 'atrasadas': 0})  # por Tipo de cobrança
for v in vendas:
    for n, (status, data_esp, valor, dpg) in v['parcelas'].items():
        if status == 'paga':
            hist[v['metodo']]['pagas'] += 1
            hist_tipo[str(v['tipo'])]['pagas'] += 1
        elif status == 'atrasada':
            hist[v['metodo']]['atrasadas'] += 1
            hist_tipo[str(v['tipo'])]['atrasadas'] += 1

def taxa_inad(d):
    tot = d['pagas'] + d['atrasadas']
    return (d['atrasadas'] / tot) if tot > 0 else 0.0

taxa_geral = (total_parc_atrasadas / total_parc_venc) if total_parc_venc else 0.0

# Taxa por método de pagamento (informativa). Base pequena (<5) usa taxa geral.
taxa_por_metodo = {}
for metodo, d in hist.items():
    tot = d['pagas'] + d['atrasadas']
    taxa_por_metodo[metodo] = taxa_inad(d) if tot >= 5 else taxa_geral

# Taxa por TIPO DE COBRANÇA — usada para AJUSTAR A PROJEÇÃO (mais preciso:
# parcelas futuras só existem em Recuperador/Parcelado Hotmart/Assinatura, com risco bem distinto).
taxa_por_tipo = {}
for tp, d in hist_tipo.items():
    tot = d['pagas'] + d['atrasadas']
    taxa_por_tipo[tp] = taxa_inad(d) if tot >= 5 else taxa_geral

# ========== Workbook ==========
wb = Workbook()

# ---------- Aba 1: Parcelas por Venda ----------
ws = wb.active
ws.title = "Parcelas por Venda"
MAX_P = 12

headers_fixos = [
    "Email", "Nome", "Código da venda", "Produto", "Tipo de plano",
    "Data da venda", "Total parcelas", "Valor líquido/parcela",
    "Pagas", "Atrasadas", "Futuras", "Desconhecidas", "Fora do período", "Status geral"
]
headers = headers_fixos + [f"P{i}" for i in range(1, MAX_P + 1)]
ws.append(headers)

for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

for v in vendas:
    parc = v['parcelas']
    pagas = sum(1 for s, _, _, _ in parc.values() if s == 'paga')
    atrasadas = sum(1 for s, _, _, _ in parc.values() if s == 'atrasada')
    futuras = sum(1 for s, _, _, _ in parc.values() if s == 'futura')
    desc = sum(1 for s, _, _, _ in parc.values() if s == 'desconhecida')
    fora = sum(1 for s, _, _, _ in parc.values() if s == 'fora_periodo')

    if atrasadas > 0:
        status_geral = "Inadimplente"
    elif desc > 0:
        status_geral = "Status incerto"
    elif futuras > 0:
        status_geral = "Em dia"
    else:
        status_geral = "Concluído"

    row = [
        v['email'], v['nome'], v['id_venda'], v['produto'], v['tipo'],
        v['data_venda'].strftime("%d/%m/%Y") if v['data_venda'] else "",
        v['total_parcelas'], round(v['valor_liquido_parcela'], 2),
        pagas, atrasadas, futuras, desc, fora, status_geral,
    ]
    for n in range(1, MAX_P + 1):
        if n in parc:
            status, _, _, _ = parc[n]
            if status == 'paga':
                row.append("X")
            elif status == 'fora_periodo':
                row.append("?")
            elif status == 'desconhecida':
                row.append("?")
            else:
                row.append("")
        else:
            row.append("")
    ws.append(row)

    r = ws.max_row
    for n in range(1, MAX_P + 1):
        col = len(headers_fixos) + n
        cell = ws.cell(row=r, column=col)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
        if n in parc:
            status, dt, _, _ = parc[n]
            if status == 'paga':
                cell.fill = FILL_PAGA
            elif status == 'atrasada':
                cell.fill = FILL_ATRASADA
            elif status == 'fora_periodo':
                cell.fill = FILL_FORA
            elif status == 'desconhecida':
                cell.fill = FILL_DESC
            else:
                cell.fill = FILL_FUTURA

    # border nas demais colunas
    for col in range(1, len(headers_fixos) + 1):
        ws.cell(row=r, column=col).border = BORDER

# Congelar painel e filtro
ws.freeze_panes = "N2"
ws.auto_filter.ref = ws.dimensions

# Larguras
larguras = [32, 28, 18, 14, 22, 13, 10, 14, 8, 11, 10, 14, 14, 16] + [5] * MAX_P
for i, w in enumerate(larguras, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30

# ---------- Aba 2: Projeção 12 meses ----------
ws2 = wb.create_sheet("Projeção 12 meses")

# Calcular meses dos próximos 12
mes_ref = datetime(HOJE.year, HOJE.month, 1)
meses = [(mes_ref + relativedelta(months=i)) for i in range(12)]
meses_lbl = [m.strftime("%m/%Y") for m in meses]

# Somar valores futuros por mês e por tipo
from collections import defaultdict
proj_recup = defaultdict(float)
proj_assin = defaultdict(float)
proj_outros = defaultdict(float)
qtd_parc_futuras = defaultdict(int)
proj_ajustada = defaultdict(float)  # já descontada a inadimplência histórica por tipo de cobrança

for v in vendas:
    taxa_v = taxa_por_tipo.get(str(v['tipo']), taxa_geral)
    for n, (status, data_esp, valor, data_pgto) in v['parcelas'].items():
        if status != 'futura':
            continue
        chave = datetime(data_esp.year, data_esp.month, 1)
        if chave not in meses:
            continue
        qtd_parc_futuras[chave] += 1
        proj_ajustada[chave] += valor * (1 - taxa_v)
        if v['tipo'] == 'Assinatura':
            proj_assin[chave] += valor
        elif v['tipo'] and 'Recuperador' in str(v['tipo']):
            proj_recup[chave] += valor
        else:
            proj_outros[chave] += valor

ws2.append(["Projeção de Faturamento Líquido - próximos 12 meses (cenário otimista: 100% das parcelas futuras pagas)"])
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
ws2.cell(row=1, column=1).font = Font(bold=True, size=12)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center")

ws2.append([])

header2 = ["Categoria"] + meses_lbl + ["Total 12m"]
ws2.append(header2)
for i in range(1, len(header2) + 1):
    c = ws2.cell(row=3, column=i)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

def add_linha(label, dados, bold=False):
    row = [label]
    total = 0
    for m in meses:
        val = dados.get(m, 0.0)
        row.append(round(val, 2))
        total += val
    row.append(round(total, 2))
    ws2.append(row)
    r = ws2.max_row
    ws2.cell(row=r, column=1).font = FONT_BOLD if bold else Font()
    for i in range(1, len(row) + 1):
        ws2.cell(row=r, column=i).border = BORDER
        if i > 1:
            ws2.cell(row=r, column=i).number_format = 'R$ #,##0.00'

add_linha("Recuperador Inteligente / Parcelado", proj_recup)
add_linha("Assinatura (Raízes - R)", proj_assin)
add_linha("Outros", proj_outros)

total_mes = {m: proj_recup.get(m, 0) + proj_assin.get(m, 0) + proj_outros.get(m, 0) for m in meses}
add_linha("TOTAL LÍQUIDO PROJETADO", total_mes, bold=True)
r_total = ws2.max_row
ws2.cell(row=r_total, column=1).fill = FILL_SUBHEADER
for i in range(2, len(meses) + 3):
    ws2.cell(row=r_total, column=i).fill = FILL_SUBHEADER
    ws2.cell(row=r_total, column=i).font = FONT_BOLD

# Linha qtd parcelas
ws2.append([])
ws2.append(["Qtd. parcelas previstas"] + [qtd_parc_futuras.get(m, 0) for m in meses] +
           [sum(qtd_parc_futuras.values())])
r = ws2.max_row
ws2.cell(row=r, column=1).font = FONT_BOLD
for i in range(1, 14):
    ws2.cell(row=r, column=i).border = BORDER

# Larguras
ws2.column_dimensions['A'].width = 40
for i in range(2, 15):
    ws2.column_dimensions[get_column_letter(i)].width = 13

# Nota
ws2.append([])
ws2.append(["Obs.: projeção considera apenas parcelas de vendas já realizadas e ainda não cobradas."])
ws2.append(["Não inclui novas vendas futuras. Assume 100% de pagamento (cenário otimista)."])
ws2.append(["Valores em R$ líquido (já descontadas taxas Hotmart)."])

# ---------- Aba: Resumo Mensal (Total + Acumulado) ----------
ws_res = wb.create_sheet("Resumo Mensal", 2)  # posição 3 (entre Projeção e Inadimplência)

ws_res.append(["RESUMO MENSAL — QUANTO TEMOS A RECEBER (LÍQUIDO)"])
ws_res.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws_res.cell(row=1, column=1).font = Font(bold=True, size=13, color="FFFFFF")
ws_res.cell(row=1, column=1).fill = FILL_HEADER
ws_res.cell(row=1, column=1).alignment = Alignment(horizontal="center")
ws_res.row_dimensions[1].height = 26

ws_res.append([f"Data de referência: {HOJE.strftime('%d/%m/%Y')} | Otimista = 100% | Ajustado = descontada inadimplência histórica por TIPO DE COBRANÇA"])
ws_res.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws_res.cell(row=2, column=1).alignment = Alignment(horizontal="center")
ws_res.cell(row=2, column=1).font = Font(italic=True, color="595959")

ws_res.append([])

hdrs = ["Mês", "Qtd. parcelas", "Otimista (100%) R$",
        "Inadimplência esperada R$", "Ajustado a receber R$", "Acumulado ajustado R$"]
ws_res.append(hdrs)
r_hdr = ws_res.max_row
for i in range(1, len(hdrs) + 1):
    c = ws_res.cell(row=r_hdr, column=i)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER
ws_res.row_dimensions[r_hdr].height = 32

acum = 0.0
for m, lbl in zip(meses, meses_lbl):
    val = total_mes.get(m, 0.0)
    val_aj = proj_ajustada.get(m, 0.0)
    inad = val - val_aj
    qtd = qtd_parc_futuras.get(m, 0)
    acum += val_aj
    ws_res.append([lbl, qtd, round(val, 2), round(inad, 2), round(val_aj, 2), round(acum, 2)])
    r = ws_res.max_row
    for i in range(1, 7):
        cell = ws_res.cell(row=r, column=i)
        cell.border = BORDER
        if i >= 3:
            cell.number_format = 'R$ #,##0.00'
        if i == 1:
            cell.alignment = Alignment(horizontal="center")
            cell.font = FONT_BOLD
    ws_res.cell(row=r, column=4).font = Font(color="C00000")

# Total
tot_otim = sum(total_mes.values())
tot_aj = sum(proj_ajustada.values())
ws_res.append(["TOTAL 12 MESES", sum(qtd_parc_futuras.values()),
               round(tot_otim, 2), round(tot_otim - tot_aj, 2),
               round(tot_aj, 2), round(tot_aj, 2)])
r = ws_res.max_row
for i in range(1, 7):
    cell = ws_res.cell(row=r, column=i)
    cell.fill = FILL_SUBHEADER
    cell.font = FONT_BOLD
    cell.border = BORDER
    if i >= 3:
        cell.number_format = 'R$ #,##0.00'

# Valor já vencido em atraso (contexto)
ws_res.append([])
ws_res.append(["REFERÊNCIA — já vencido (não está na projeção acima):"])
r = ws_res.max_row
ws_res.cell(row=r, column=1).font = FONT_BOLD
ws_res.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

for label, valor, fill in [
    ("Valor em ATRASO confirmado (Hotmart tentou cobrar e falhou)", valor_em_atraso, FILL_ATRASADA),
    ("Valor em situação DESCONHECIDA (sem registro, possível cancelamento)", valor_desc, FILL_DESC),
]:
    ws_res.append([label, "", round(valor, 2), "", "", ""])
    r = ws_res.max_row
    ws_res.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for i in range(1, 7):
        c = ws_res.cell(row=r, column=i)
        c.border = BORDER
        c.fill = fill
        if i == 3:
            c.number_format = 'R$ #,##0.00'
            c.font = FONT_BOLD

# Larguras
ws_res.column_dimensions['A'].width = 58
ws_res.column_dimensions['B'].width = 13
for col in ['C', 'D', 'E', 'F']:
    ws_res.column_dimensions[col].width = 20

# Gráfico de barras: Otimista vs Ajustado
from openpyxl.chart import BarChart, Reference, LineChart
chart = BarChart()
chart.type = "col"
chart.title = "A receber por mês: Otimista vs Ajustado (R$ líquido)"
chart.y_axis.title = "R$"
chart.x_axis.title = "Mês"
data = Reference(ws_res, min_col=3, max_col=3, min_row=4, max_row=4 + len(meses))
data_aj = Reference(ws_res, min_col=5, max_col=5, min_row=4, max_row=4 + len(meses))
chart.add_data(data, titles_from_data=True)
chart.add_data(data_aj, titles_from_data=True)
cats = Reference(ws_res, min_col=1, min_row=5, max_row=4 + len(meses))
chart.set_categories(cats)
chart.height = 10
chart.width = 24
ws_res.add_chart(chart, "H4")

# ---------- Aba: Histórico Recebido (passado) ----------
ws_hist = wb.create_sheet("Histórico Recebido", 3)

# Somar valores de parcelas PAGAS, agrupado por mês de recebimento e tipo
from collections import defaultdict
hist_recup = defaultdict(float)
hist_assin = defaultdict(float)
hist_avista = defaultdict(float)
qtd_pagas_mes = defaultdict(int)
hist_meses = set()

for v in vendas:
    tipo = v['tipo'] or ''
    for n, (status, data_esp, valor, data_pgto) in v['parcelas'].items():
        if status != 'paga' or data_pgto is None:
            continue
        chave = datetime(data_pgto.year, data_pgto.month, 1)
        if chave > datetime(HOJE.year, HOJE.month, 1):
            continue
        hist_meses.add(chave)
        qtd_pagas_mes[chave] += 1
        if tipo == 'Assinatura':
            hist_assin[chave] += valor
        elif 'Recuperador' in tipo or 'Parcelado' in tipo or 'Parcelamento Inteligente' in tipo:
            hist_recup[chave] += valor
        else:
            hist_avista[chave] += valor

meses_hist = sorted(hist_meses)
# Limitar aos últimos ~12 meses
if len(meses_hist) > 12:
    meses_hist = meses_hist[-12:]

ws_hist.append(["HISTÓRICO DE RECEBIMENTO MENSAL (LÍQUIDO)"])
ws_hist.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
ws_hist.cell(row=1, column=1).font = Font(bold=True, size=13, color="FFFFFF")
ws_hist.cell(row=1, column=1).fill = FILL_HEADER
ws_hist.cell(row=1, column=1).alignment = Alignment(horizontal="center")
ws_hist.row_dimensions[1].height = 26

ws_hist.append([f"Período: {meses_hist[0].strftime('%m/%Y')} a {meses_hist[-1].strftime('%m/%Y')} | Data ref: {HOJE.strftime('%d/%m/%Y')}"])
ws_hist.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws_hist.cell(row=2, column=1).alignment = Alignment(horizontal="center")
ws_hist.cell(row=2, column=1).font = Font(italic=True, color="595959")

ws_hist.append([])

hdrs_h = ["Mês", "Qtd. parcelas pagas", "Parcelado/Recuperador (R$)",
          "Assinatura (R$)", "À vista (R$)", "Total recebido (R$)", "Acumulado (R$)"]
ws_hist.append(hdrs_h)
r_hdr = ws_hist.max_row
for i in range(1, len(hdrs_h) + 1):
    c = ws_hist.cell(row=r_hdr, column=i)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

acum_h = 0.0
for m in meses_hist:
    rec = hist_recup.get(m, 0.0)
    ass = hist_assin.get(m, 0.0)
    av = hist_avista.get(m, 0.0)
    tot = rec + ass + av
    acum_h += tot
    ws_hist.append([
        m.strftime("%m/%Y"), qtd_pagas_mes.get(m, 0),
        round(rec, 2), round(ass, 2), round(av, 2),
        round(tot, 2), round(acum_h, 2),
    ])
    r = ws_hist.max_row
    for i in range(1, 8):
        cell = ws_hist.cell(row=r, column=i)
        cell.border = BORDER
        if i >= 3:
            cell.number_format = 'R$ #,##0.00'
        if i == 1:
            cell.alignment = Alignment(horizontal="center")
            cell.font = FONT_BOLD

# Total geral
total_rec = sum(hist_recup.values())
total_ass = sum(hist_assin.values())
total_av = sum(hist_avista.values())
total_g = total_rec + total_ass + total_av
ws_hist.append([
    "TOTAL", sum(qtd_pagas_mes.get(m, 0) for m in meses_hist),
    round(total_rec, 2), round(total_ass, 2), round(total_av, 2),
    round(total_g, 2), round(total_g, 2),
])
r = ws_hist.max_row
for i in range(1, 8):
    cell = ws_hist.cell(row=r, column=i)
    cell.fill = FILL_SUBHEADER
    cell.font = FONT_BOLD
    cell.border = BORDER
    if i >= 3:
        cell.number_format = 'R$ #,##0.00'

# Média mensal
ws_hist.append([])
ws_hist.append([
    "Média mensal", round(sum(qtd_pagas_mes.get(m, 0) for m in meses_hist) / len(meses_hist), 1),
    round(total_rec / len(meses_hist), 2),
    round(total_ass / len(meses_hist), 2),
    round(total_av / len(meses_hist), 2),
    round(total_g / len(meses_hist), 2),
    "",
])
r = ws_hist.max_row
for i in range(1, 8):
    cell = ws_hist.cell(row=r, column=i)
    cell.border = BORDER
    cell.font = FONT_BOLD
    if i >= 3 and i <= 6:
        cell.number_format = 'R$ #,##0.00'

# Larguras
ws_hist.column_dimensions['A'].width = 12
ws_hist.column_dimensions['B'].width = 18
ws_hist.column_dimensions['C'].width = 24
ws_hist.column_dimensions['D'].width = 18
ws_hist.column_dimensions['E'].width = 18
ws_hist.column_dimensions['F'].width = 22
ws_hist.column_dimensions['G'].width = 22
ws_hist.row_dimensions[4].height = 32

# Gráfico empilhado
from openpyxl.chart import BarChart as _BC
chart_h = _BC()
chart_h.type = "col"
chart_h.grouping = "stacked"
chart_h.overlap = 100
chart_h.title = "Recebido por mês (R$ líquido) — empilhado por tipo"
chart_h.y_axis.title = "R$"
chart_h.x_axis.title = "Mês"
data_h = Reference(ws_hist, min_col=3, max_col=5, min_row=4, max_row=4 + len(meses_hist))
cats_h = Reference(ws_hist, min_col=1, min_row=5, max_row=4 + len(meses_hist))
chart_h.add_data(data_h, titles_from_data=True)
chart_h.set_categories(cats_h)
chart_h.height = 10
chart_h.width = 22
ws_hist.add_chart(chart_h, "I4")

# ---------- Aba: Devedores por Cliente ----------
ws_dev = wb.create_sheet("Devedores por Cliente", 4)

# Agrega por cliente (email): consolida várias vendas do mesmo cliente
dev = {}
for v in vendas:
    email = v['email'] or '(sem email)'
    atrasadas_v = [(n, val) for n, (s, de, val, dp) in v['parcelas'].items() if s == 'atrasada']
    d = dev.setdefault(email, {
        'nome': v['nome'], 'vendas': 0, 'vendas_inad': 0,
        'parc_atras': 0, 'valor_atraso': 0.0, 'tipos': set(), 'metodos': set()})
    d['vendas'] += 1
    if atrasadas_v:
        d['vendas_inad'] += 1
        d['parc_atras'] += len(atrasadas_v)
        d['valor_atraso'] += sum(val for _, val in atrasadas_v)
        d['tipos'].add(str(v['tipo']))
        d['metodos'].add(v['metodo'])

devedores = [(e, d) for e, d in dev.items() if d['vendas_inad'] > 0]
devedores.sort(key=lambda x: -x[1]['valor_atraso'])

ws_dev.append(["DEVEDORES POR CLIENTE (consolidado por e-mail — várias vendas do mesmo cliente somadas)"])
ws_dev.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
ws_dev.cell(row=1, column=1).font = Font(bold=True, size=13, color="FFFFFF")
ws_dev.cell(row=1, column=1).fill = FILL_HEADER
ws_dev.cell(row=1, column=1).alignment = Alignment(horizontal="center")
ws_dev.row_dimensions[1].height = 26
ws_dev.append([f"Data ref: {HOJE.strftime('%d/%m/%Y')} | Devedor = cliente com ≥1 parcela vencida não recebida (tentativas repetidas da Hotmart contadas 1×)"])
ws_dev.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
ws_dev.cell(row=2, column=1).font = Font(italic=True, color="595959")
ws_dev.append([])

hdr_d = ["E-mail", "Nome", "Qtd vendas", "Vendas inadimplentes",
         "Parcelas atrasadas", "Valor em atraso (líq. R$)", "Tipos de cobrança", "Métodos de pagamento"]
ws_dev.append(hdr_d)
rh = ws_dev.max_row
for i in range(1, len(hdr_d) + 1):
    c = ws_dev.cell(row=rh, column=i)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER
ws_dev.row_dimensions[rh].height = 30

tot_val = 0.0
for email, d in devedores:
    ws_dev.append([
        email, d['nome'], d['vendas'], d['vendas_inad'],
        d['parc_atras'], round(d['valor_atraso'], 2),
        ", ".join(sorted(d['tipos'])), ", ".join(sorted(d['metodos']))])
    r = ws_dev.max_row
    tot_val += d['valor_atraso']
    for i in range(1, len(hdr_d) + 1):
        cell = ws_dev.cell(row=r, column=i)
        cell.border = BORDER
        if i == 6:
            cell.number_format = 'R$ #,##0.00'
    ws_dev.cell(row=r, column=6).fill = FILL_ATRASADA

ws_dev.append([])
ws_dev.append(["TOTAL", f"{len(devedores)} clientes devedores", "", "",
               sum(d['parc_atras'] for _, d in devedores), round(tot_val, 2), "", ""])
r = ws_dev.max_row
for i in range(1, len(hdr_d) + 1):
    ws_dev.cell(row=r, column=i).font = FONT_BOLD
    ws_dev.cell(row=r, column=i).fill = FILL_SUBHEADER
    if i == 6:
        ws_dev.cell(row=r, column=i).number_format = 'R$ #,##0.00'

ws_dev.freeze_panes = "A5"
ws_dev.auto_filter.ref = f"A4:H{4 + len(devedores)}"
for col, w in zip("ABCDEFGH", [34, 26, 11, 14, 13, 18, 34, 30]):
    ws_dev.column_dimensions[col].width = w

# ---------- Aba 4: Inadimplência ----------
ws3 = wb.create_sheet("Inadimplência")

def pct(a, b):
    return (a / b * 100) if b else 0

ws3.append(["RESUMO DE INADIMPLÊNCIA"])
ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
ws3.append([f"Data de referência: {HOJE.strftime('%d/%m/%Y')}"])
ws3.append([])

ws3.append(["Métrica", "Valor"])
for i in (1, 2):
    ws3.cell(row=4, column=i).fill = FILL_HEADER
    ws3.cell(row=4, column=i).font = FONT_HEADER
    ws3.cell(row=4, column=i).border = BORDER

linhas = [
    ("Total de vendas (contratos)", total_vendas),
    ("Total de alunos únicos (emails)", len(emails_unicos)),
    ("Vendas com pelo menos 1 parcela em atraso CONFIRMADO", vendas_com_atraso),
    ("% vendas inadimplentes (confirmado)", f"{pct(vendas_com_atraso, total_vendas):.1f}%"),
    ("Alunos únicos inadimplentes (confirmado)", len(emails_inad)),
    ("% alunos inadimplentes (confirmado)", f"{pct(len(emails_inad), len(emails_unicos)):.1f}%"),
    ("", ""),
    ("--- PARCELAS (apenas as que têm registro no CSV) ---", ""),
    ("Parcelas PAGAS (Completo)", total_parc_pagas),
    ("Parcelas ATRASADAS (Hotmart tentou e falhou)", total_parc_atrasadas),
    ("Parcelas DESCONHECIDAS (data passou mas sem registro no CSV)", total_parc_desc),
    ("% inadimplência entre parcelas com registro (atrasadas / (pagas+atrasadas))",
     f"{pct(total_parc_atrasadas, total_parc_venc):.1f}%"),
    ("", ""),
    ("--- VALORES LÍQUIDOS ---", ""),
    ("Valor já recebido (parcelas pagas)", f"R$ {valor_recebido:,.2f}"),
    ("Valor em atraso CONFIRMADO (tentou cobrar e falhou)", f"R$ {valor_em_atraso:,.2f}"),
    ("Valor em situação DESCONHECIDA (possível cancelamento)", f"R$ {valor_desc:,.2f}"),
    ("% valor em atraso confirmado / (recebido + atrasado)",
     f"{pct(valor_em_atraso, valor_recebido + valor_em_atraso):.1f}%"),
]
for lbl, val in linhas:
    ws3.append([lbl, val])
    r = ws3.max_row
    ws3.cell(row=r, column=1).font = FONT_BOLD
    ws3.cell(row=r, column=1).border = BORDER
    ws3.cell(row=r, column=2).border = BORDER

# Quebra por tipo de plano
ws3.append([])
ws3.append(["QUEBRA POR TIPO DE PLANO"])
r = ws3.max_row
ws3.cell(row=r, column=1).font = Font(bold=True, size=12)
ws3.append([])

headers_tipo = ["Tipo de plano", "Qtd. vendas", "Vendas inad.", "% vendas inad.",
                "Parc. pagas", "Parc. atrasadas", "Parc. desconhec.",
                "Valor em atraso (líq.)"]
ws3.append(headers_tipo)
r_hdr = ws3.max_row
for i in range(1, len(headers_tipo) + 1):
    ws3.cell(row=r_hdr, column=i).fill = FILL_HEADER
    ws3.cell(row=r_hdr, column=i).font = FONT_HEADER
    ws3.cell(row=r_hdr, column=i).border = BORDER
    ws3.cell(row=r_hdr, column=i).alignment = Alignment(horizontal="center", wrap_text=True)

tipos = sorted(set((v['tipo'] or 'Sem tipo') for v in vendas))
for tipo in tipos:
    vs = [v for v in vendas if (v['tipo'] or 'Sem tipo') == tipo]
    qt = len(vs)
    inad = sum(1 for v in vs if any(s == 'atrasada' for s, _, _, _ in v['parcelas'].values()))
    pp = 0; pa = 0; pd = 0; va = 0.0
    for v in vs:
        for n, (status, _, valor, _) in v['parcelas'].items():
            if status == 'paga':
                pp += 1
            elif status == 'atrasada':
                pa += 1
                va += valor
            elif status == 'desconhecida':
                pd += 1
    ws3.append([
        tipo, qt, inad, f"{pct(inad, qt):.1f}%",
        pp, pa, pd, f"R$ {va:,.2f}"
    ])
    r = ws3.max_row
    for i in range(1, len(headers_tipo) + 1):
        ws3.cell(row=r, column=i).border = BORDER

ws3.column_dimensions['A'].width = 55
for i in range(2, 9):
    ws3.column_dimensions[get_column_letter(i)].width = 18

# ===== Tabela: % inadimplência histórica por FORMA DE PAGAMENTO =====
ws3.append([])
ws3.append([])
ws3.append(["INADIMPLÊNCIA HISTÓRICA POR FORMA DE PAGAMENTO (base p/ ajustar a projeção)"])
r = ws3.max_row
ws3.cell(row=r, column=1).font = Font(bold=True, size=12)
ws3.append(["Base: parcelas já vencidas (pagas + atrasadas confirmadas). Métodos com <5 parcelas usam a taxa geral na projeção."])
ws3.cell(row=ws3.max_row, column=1).font = Font(italic=True, color="595959")
ws3.append([])

hdr_fp = ["Forma de pagamento", "Parcelas pagas", "Parcelas atrasadas",
          "Total vencidas", "% inadimplência", "Taxa usada na projeção"]
ws3.append(hdr_fp)
r_h = ws3.max_row
for i in range(1, len(hdr_fp) + 1):
    c = ws3.cell(row=r_h, column=i)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

for metodo, d in sorted(hist.items(), key=lambda x: -taxa_inad(x[1])):
    tot = d['pagas'] + d['atrasadas']
    if tot == 0:
        continue
    taxa_real = taxa_inad(d) * 100
    taxa_proj = taxa_por_metodo.get(metodo, taxa_geral) * 100
    ws3.append([metodo, d['pagas'], d['atrasadas'], tot,
                f"{taxa_real:.1f}%", f"{taxa_proj:.1f}%"])
    rr = ws3.max_row
    for i in range(1, len(hdr_fp) + 1):
        ws3.cell(row=rr, column=i).border = BORDER
    ws3.cell(row=rr, column=5).font = Font(bold=True, color="C00000")

ws3.append(["GERAL (todas as formas)", total_parc_pagas, total_parc_atrasadas,
            total_parc_venc, f"{taxa_geral*100:.1f}%", f"{taxa_geral*100:.1f}%"])
rr = ws3.max_row
for i in range(1, len(hdr_fp) + 1):
    ws3.cell(row=rr, column=i).fill = FILL_SUBHEADER
    ws3.cell(row=rr, column=i).font = FONT_BOLD
    ws3.cell(row=rr, column=i).border = BORDER

# ===== Tabela: % inadimplência histórica por TIPO DE COBRANÇA (usada na projeção) =====
ws3.append([])
ws3.append([])
ws3.append(["INADIMPLÊNCIA HISTÓRICA POR TIPO DE COBRANÇA  ← esta é a taxa usada para ajustar a projeção"])
r = ws3.max_row
ws3.cell(row=r, column=1).font = Font(bold=True, size=12, color="C00000")
ws3.append(["Mais preciso que método de pagamento: parcelas futuras só existem em Recuperador/Parcelado Hotmart/Assinatura, com riscos distintos."])
ws3.cell(row=ws3.max_row, column=1).font = Font(italic=True, color="595959")
ws3.append([])

hdr_tp = ["Tipo de cobrança", "Parcelas pagas", "Parcelas não recebidas",
          "Total vencidas", "% inadimplência", "Taxa usada na projeção"]
ws3.append(hdr_tp)
r_h = ws3.max_row
for i in range(1, len(hdr_tp) + 1):
    c = ws3.cell(row=r_h, column=i)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

for tp, d in sorted(hist_tipo.items(), key=lambda x: -taxa_inad(x[1])):
    tot = d['pagas'] + d['atrasadas']
    if tot == 0:
        continue
    ws3.append([tp, d['pagas'], d['atrasadas'], tot,
                f"{taxa_inad(d)*100:.1f}%", f"{taxa_por_tipo.get(tp, taxa_geral)*100:.1f}%"])
    rr = ws3.max_row
    for i in range(1, len(hdr_tp) + 1):
        ws3.cell(row=rr, column=i).border = BORDER
    ws3.cell(row=rr, column=5).font = Font(bold=True, color="C00000")

ws3.append(["GERAL", total_parc_pagas, total_parc_atrasadas,
            total_parc_venc, f"{taxa_geral*100:.1f}%", f"{taxa_geral*100:.1f}%"])
rr = ws3.max_row
for i in range(1, len(hdr_tp) + 1):
    ws3.cell(row=rr, column=i).fill = FILL_SUBHEADER
    ws3.cell(row=rr, column=i).font = FONT_BOLD
    ws3.cell(row=rr, column=i).border = BORDER

# Legenda na aba 1
ws.append([])
ws.append([])
leg_row = ws.max_row + 1
ws.cell(row=leg_row, column=1, value="LEGENDA:").font = FONT_BOLD
ws.cell(row=leg_row, column=2, value="X = paga").fill = FILL_PAGA
ws.cell(row=leg_row, column=3, value="(vazio vermelho) = atrasada (Hotmart tentou cobrar e falhou)").fill = FILL_ATRASADA
ws.cell(row=leg_row, column=4, value="(vazio cinza) = futura (data ainda não chegou)").fill = FILL_FUTURA
ws.cell(row=leg_row, column=5, value="? laranja = DESCONHECIDA (data passou mas sem registro - possível cancelamento)").fill = FILL_DESC
ws.cell(row=leg_row, column=6, value="? amarelo = ANTES do período do CSV").fill = FILL_FORA

print(f"Salvando em {OUT}...")
try:
    wb.save(OUT)
except PermissionError:
    _alt = OUT.replace('.xlsx', datetime.now().strftime('_%Y%m%d_%H%M.xlsx'))
    wb.save(_alt)
    OUT = _alt
    print(f"  [!] {OUT.split(chr(92))[-1]} estava aberto no Excel. Salvei como: {_alt.split(chr(92))[-1]}")
print("OK!")
print(f"  - {total_vendas} vendas")
print(f"  - {len(emails_unicos)} alunos únicos")
print(f"  - {vendas_com_atraso} vendas com atraso confirmado ({pct(vendas_com_atraso, total_vendas):.1f}%)")
print(f"  - Valor em atraso CONFIRMADO (líq.): R$ {valor_em_atraso:,.2f}")
print(f"  - Valor DESCONHECIDO (líq.): R$ {valor_desc:,.2f}")
print(f"  - Projeção 12m OTIMISTA: R$ {sum(total_mes.values()):,.2f}")
print(f"  - Projeção 12m AJUSTADA (- inadimplência): R$ {sum(proj_ajustada.values()):,.2f}")
print(f"  - Taxa geral de inadimplência histórica: {taxa_geral*100:.1f}%")

# ============================================================
#  DASHBOARD HTML DINÂMICO (gerado com os números recalculados)
# ============================================================
import json as _json

OUT_DASH = OUT.replace('.xlsx', '_dashboard.html').replace('base_mare', 'dashboard_mare') \
    if 'base_mare' in OUT else OUT.replace('.xlsx', '_dashboard.html')
OUT_DASH = r"C:\Users\saram\Downloads\dashboard_mare.html"

# --- Série mensal da projeção ---
_meses_lbl = [m.strftime('%m/%y') for m in meses]
_otim = [round(total_mes.get(m, 0)) for m in meses]
_ajus = [round(proj_ajustada.get(m, 0)) for m in meses]
_qtd = [int(qtd_parc_futuras.get(m, 0)) for m in meses]
_ac = []
_s = 0
for v in _ajus:
    _s += v
    _ac.append(_s)

# --- Histórico recebido ---
_hist_lbl = [m.strftime('%m/%y') for m in meses_hist]
_hist_tot = [round(hist_recup.get(m, 0) + hist_assin.get(m, 0) + hist_avista.get(m, 0)) for m in meses_hist]

# --- Inadimplência por tipo de cobrança ---
_it = sorted([(t, taxa_inad(d)) for t, d in hist_tipo.items()
              if (d['pagas'] + d['atrasadas']) > 0], key=lambda x: -x[1])
_it_lbl = [t for t, _ in _it]
_it_val = [round(p * 100, 1) for _, p in _it]

# --- Inadimplência por forma de pagamento ---
_ip = sorted([(t, taxa_inad(d), d['pagas'] + d['atrasadas']) for t, d in hist.items()
              if (d['pagas'] + d['atrasadas']) >= 5], key=lambda x: -x[1])
_ip_lbl = [t for t, _, _ in _ip]
_ip_val = [round(p * 100, 1) for _, p, _ in _ip]

# --- Assinatura por método ---
_assin = {}
for v in vendas:
    if str(v['tipo']) != 'Assinatura':
        continue
    a = _assin.setdefault(v['metodo'], {'pg': 0, 'at': 0, 'vd': 0})
    a['vd'] += 1
    for n, (s, de, val, dp) in v['parcelas'].items():
        if s == 'paga':
            a['pg'] += 1
        elif s == 'atrasada':
            a['at'] += 1
_assin_l = sorted([(k, d) for k, d in _assin.items() if (d['pg'] + d['at']) > 0],
                  key=lambda x: -(x[1]['vd']))
_assin_lbl = [f"{k} ({d['vd']})" for k, d in _assin_l]
_assin_val = [round(d['at'] / (d['pg'] + d['at']) * 100, 1) for _, d in _assin_l]

# --- Quebra por tipo de plano ---
_qt = {}
for v in vendas:
    t = str(v['tipo'])
    q = _qt.setdefault(t, {'vd': 0, 'dev': 0, 'val': 0.0})
    q['vd'] += 1
    atr = [val for n, (s, de, val, dp) in v['parcelas'].items() if s == 'atrasada']
    if atr:
        q['dev'] += 1
        q['val'] += sum(atr)
_qt_rows = sorted(_qt.items(), key=lambda x: -x[1]['val'])
_qt_data = [[t, d['vd'], d['dev'],
             f"{(d['dev']/d['vd']*100 if d['vd'] else 0):.1f}%", round(d['val'], 2)]
            for t, d in _qt_rows if d['vd'] > 0]

# --- Top 12 devedores ---
_top = [[d['nome'][:32], d['parc_atras'], round(d['valor_atraso'], 2)]
        for e, d in devedores[:12]]

# --- Distribuição devedores por nº de parcelas em atraso ---
_distc, _distv = {}, {}
for e, d in devedores:
    n = d['parc_atras']
    _distc[n] = _distc.get(n, 0) + 1
    _distv[n] = _distv.get(n, 0.0) + d['valor_atraso']
_dn = sorted(_distc)
_dist_lbl = [str(n) for n in _dn]
_dist_cli = [_distc[n] for n in _dn]
_dist_val = [round(_distv[n]) for n in _dn]

# --- Lançamentos: por mês de venda, separa Total / Parcelado / Inadimplência sobre parcelado ---
_TIPOS_RISCO = {'Recuperador Inteligente', 'Assinatura',
                'Parcelado Hotmart - Recuperador', 'Parcelado Hotmart - Link exclusivo',
                'Parcelado Hotmart - Checkout', 'Parcelamento Inteligente'}
_lanc = {}
for v in vendas:
    mes = v['data_venda'].strftime('%m/%Y')
    risco = str(v['tipo']) in _TIPOS_RISCO
    inad = any(s == 'atrasada' for s, _, _, _ in v['parcelas'].values())
    a = _lanc.setdefault(mes, {'total': 0, 'risco': 0, 'inad_risco': 0, 'ord': v['data_venda']})
    a['total'] += 1
    if risco:
        a['risco'] += 1
        if inad:
            a['inad_risco'] += 1
# Filtra meses com >=50 vendas (lançamentos) e ordena cronologicamente
_lanc_ord = sorted([(m, a) for m, a in _lanc.items() if a['total'] >= 50], key=lambda x: x[1]['ord'])
_lanc_lbl = [m for m, _ in _lanc_ord]
_lanc_total = [a['total'] for _, a in _lanc_ord]
_lanc_parc = [a['risco'] for _, a in _lanc_ord]
_lanc_inad = [a['inad_risco'] for _, a in _lanc_ord]
_lanc_pct = [round(a['inad_risco'] / a['risco'] * 100, 1) if a['risco'] else 0 for _, a in _lanc_ord]
# Rótulos com tipo de lançamento
_lanc_tipos = {'06/2025': 'Gratuito', '09/2025': 'Gratuito', '02/2026': 'Gratuito',
               '07/2025': 'Desafio 1', '10/2025': 'Desafio 2',
               '03/2026': 'Desafio 3', '05/2026': 'Desafio 4',
               '06/2026': 'Desafio 5', '04/2026': 'Desafio 4 pré'}
_lanc_rot = [_lanc_tipos.get(m, '') for m in _lanc_lbl]

# --- Análise por vendedor (SRC) ---
import re as _re
_PATS_VENDEDORES = {
    'Sara': [r'^sara'], 'Vanessa': [r'vanessa', r'vansanches'], 'Luiggi': [r'luiggi'],
    'Patricia': [r'\bpaty\b', r'jhtpaty', r'\bpatricia\b'],
    'Guilherme': [r'jhtgui', r'^gui$'],
    'Lais': [r'\blais\b'], 'Iza': [r'^iza$', r'izabela'],
    'Ana Rita': [r'anarita'], 'Vitor': [r'\bvitor\b'],
    'Priscila': [r'priscila'], 'Fernanda': [r'fernanda']
}
_AUTO_KW = ['dm7', 'grupo', 'vip', 'zoom', 'api', 'organic']

def _classifica_src(s):
    s = str(s).strip().lower() if s else ''
    if not s or s in ('(none)', 'nan', ''):
        return 'SEM SRC'
    for nome, lst in _PATS_VENDEDORES.items():
        for p in lst:
            if _re.search(p, s):
                return nome
    if any(k in s for k in _AUTO_KW):
        return 'Automático'
    return 'Outro'

# Mapa: ID_venda -> Código SRC (da linha mais antiga)
_srcs = df.sort_values('Data da transação').drop_duplicates('ID_venda', keep='first').set_index('ID_venda')['Código SRC']

_vend_agg = {}
for v in vendas:
    srcv = _srcs.get(v['id_venda'], '')
    origem = _classifica_src(srcv)
    rec = atr = fut = 0.0
    for n, (st, de, val, dp) in v['parcelas'].items():
        if st == 'paga':
            rec += val
        elif st == 'atrasada':
            atr += val
        elif st == 'futura':
            fut += val
    a = _vend_agg.setdefault(origem, {'vd': 0, 'rec': 0.0, 'atr': 0.0, 'fut': 0.0})
    a['vd'] += 1
    a['rec'] += rec
    a['atr'] += atr
    a['fut'] += fut

# Ordena por % perda desc; coloca SEM SRC e Automático no fim
_vend_ord = sorted([k for k in _vend_agg if k not in ('SEM SRC', 'Automático', 'Outro')],
                   key=lambda k: -((_vend_agg[k]['atr'] / (_vend_agg[k]['rec'] + _vend_agg[k]['atr'])) if (_vend_agg[k]['rec'] + _vend_agg[k]['atr']) else 0))
for x in ['Vanessa', 'Ana Rita', 'Automático', 'SEM SRC', 'Outro']:
    if x in _vend_agg and x not in _vend_ord:
        _vend_ord.append(x)

_vend_data = []
for k in _vend_ord:
    a = _vend_agg[k]
    venc = a['rec'] + a['atr']
    perda = a['atr'] / venc * 100 if venc else 0
    _vend_data.append([k, a['vd'], round(a['rec']), round(a['atr']), round(a['fut']),
                       round(perda, 1), round(100 - perda, 1)])

_kpi = {
    'recebido': round(valor_recebido), 'pagas': total_parc_pagas,
    'otim': round(sum(total_mes.values())), 'qfut': int(sum(qtd_parc_futuras.values())),
    'ajus': round(sum(proj_ajustada.values())),
    'atraso': round(valor_em_atraso), 'qatraso': total_parc_atrasadas,
    'taxa': round(taxa_geral * 100, 1), 'devedores': len(devedores),
    'alunos': len(emails_unicos), 'pctdev': round(len(devedores) / len(emails_unicos) * 100, 1) if emails_unicos else 0,
}

D = {
    'ref': HOJE.strftime('%d/%m/%Y'),
    'kpi': _kpi,
    'mLbl': _meses_lbl, 'otim': _otim, 'ajus': _ajus, 'qtd': _qtd, 'ac': _ac,
    'hLbl': _hist_lbl, 'hTot': _hist_tot,
    'itLbl': _it_lbl, 'itVal': _it_val,
    'ipLbl': _ip_lbl, 'ipVal': _ip_val,
    'asLbl': _assin_lbl, 'asVal': _assin_val,
    'qt': _qt_data, 'top': _top,
    'dLbl': _dist_lbl, 'dCli': _dist_cli, 'dVal': _dist_val,
    'lLbl': _lanc_lbl, 'lRot': _lanc_rot, 'lTot': _lanc_total,
    'lParc': _lanc_parc, 'lInad': _lanc_inad, 'lPct': _lanc_pct,
    'vend': _vend_data,
}

_DASH_TMPL = r'''<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Maré Educação — Dashboard Financeiro Raízes</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box;font-family:'Segoe UI',system-ui,sans-serif}
body{background:#0f172a;color:#e2e8f0;padding:24px}.wrap{max-width:1400px;margin:0 auto}
header{background:linear-gradient(135deg,#1e3a8a,#0ea5e9);border-radius:18px;padding:32px 36px;margin-bottom:24px;box-shadow:0 10px 30px rgba(14,165,233,.25)}
header h1{font-size:28px;color:#fff;margin-bottom:6px}header p{color:#bae6fd;font-size:14px}
.grid{display:grid;gap:18px}.kpis{grid-template-columns:repeat(auto-fit,minmax(200px,1fr));margin-bottom:24px}
.kpi{background:#1e293b;border-radius:14px;padding:22px;border-left:4px solid #38bdf8}
.kpi .label{font-size:12px;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}
.kpi .val{font-size:26px;font-weight:700;color:#f8fafc}.kpi .sub{font-size:12px;color:#64748b;margin-top:4px}
.kpi.green{border-color:#22c55e}.kpi.amber{border-color:#f59e0b}.kpi.red{border-color:#ef4444}.kpi.violet{border-color:#a78bfa}
.cards{grid-template-columns:repeat(12,1fr)}.card{background:#1e293b;border-radius:16px;padding:24px}
.card h2{font-size:16px;color:#f1f5f9;margin-bottom:4px}.card .desc{font-size:12px;color:#94a3b8;margin-bottom:18px}
.col-12{grid-column:span 12}.col-8{grid-column:span 8}.col-7{grid-column:span 7}.col-6{grid-column:span 6}.col-5{grid-column:span 5}.col-4{grid-column:span 4}
@media(max-width:1000px){.col-8,.col-7,.col-6,.col-5,.col-4{grid-column:span 12}}
canvas{max-height:340px}table{width:100%;border-collapse:collapse;font-size:13px}
th,td{padding:9px 12px;text-align:left;border-bottom:1px solid #334155}th{color:#94a3b8;font-size:11px;text-transform:uppercase}
td{color:#cbd5e1}tr:hover td{background:#0f172a}.num{font-variant-numeric:tabular-nums;font-weight:600}
.tag{display:inline-block;padding:3px 9px;border-radius:20px;font-size:11px;font-weight:600}
.t-hi{background:#7f1d1d;color:#fecaca}.t-md{background:#78350f;color:#fde68a}.t-lo{background:#14532d;color:#bbf7d0}
.foot{text-align:center;color:#64748b;font-size:12px;margin-top:30px;padding:18px}
.filterbar{display:flex;flex-wrap:wrap;align-items:center;gap:14px;background:#0f172a;border:1px solid #334155;border-radius:12px;padding:14px 18px}
.filterbar label{font-size:12px;color:#94a3b8;text-transform:uppercase;margin-right:6px}
.filterbar select{background:#1e293b;color:#e2e8f0;border:1px solid #475569;border-radius:8px;padding:8px 12px;font-size:14px;cursor:pointer}
.filterbar button{background:#334155;color:#cbd5e1;border:none;border-radius:8px;padding:8px 14px;font-size:13px;cursor:pointer}
.filterbar button:hover{background:#475569;color:#fff}.rangepill{margin-left:auto;display:flex;gap:18px;flex-wrap:wrap}
.rangepill div{font-size:12px;color:#94a3b8}.rangepill b{display:block;font-size:18px;color:#f8fafc;font-variant-numeric:tabular-nums}
.rangepill .g{color:#a78bfa}.rangepill .o{color:#38bdf8}
.bdg{background:#1e3a8a;color:#bae6fd;padding:2px 8px;border-radius:10px;font-size:11px;vertical-align:middle}
</style></head><body><div class="wrap">
<header><h1>📊 Maré Educação — Dashboard Financeiro Raízes</h1>
<p>Produtos: Raízes + Raízes - R · Dados até __REF__ · Valores líquidos · Gerado automaticamente</p></header>
<div class="grid kpis" id="kpis"></div>
<div class="card col-12" style="margin-bottom:6px">
<div style="font-size:12px;color:#94a3b8;margin-bottom:10px">🔎 <b style="color:#cbd5e1">Filtro de período</b> — afeta só os gráficos com <span class="bdg">filtrável</span> (Projeção e Acumulado, mensais futuros). Os demais são histórico ou foto atual.</div>
<div class="filterbar"><span><label>De</label><select id="fDe"></select></span>
<span><label>Até</label><select id="fAte"></select></span><button id="fReset">↺ Tudo</button>
<div class="rangepill"><div>Otimista no período<b class="o" id="rOt">—</b></div>
<div>Ajustado no período<b class="g" id="rAj">—</b></div><div>Parcelas<b id="rQt">—</b></div></div></div></div>
<div class="grid cards">
<div class="card col-8"><h2>Projeção mensal a receber — Otimista vs Ajustado <span class="bdg">🔎 filtrável</span></h2>
<div class="desc">Otimista = 100% · Ajustado = −inadimplência histórica por tipo de cobrança</div><canvas id="proj"></canvas></div>
<div class="card col-4"><h2>Acumulado ajustado <span class="bdg">🔎 filtrável</span></h2>
<div class="desc">Soma acumulada no período selecionado</div><canvas id="acum"></canvas></div>
<div class="card col-7"><h2>Histórico recebido mensal</h2><div class="desc">Entradas líquidas</div><canvas id="hist"></canvas></div>
<div class="card col-5"><h2>Inadimplência por tipo de cobrança</h2><div class="desc">% parcelas vencidas não recebidas (taxa usada na projeção)</div><canvas id="inadTipo"></canvas></div>
<div class="card col-6"><h2>Inadimplência por forma de pagamento</h2><div class="desc">% parcelas vencidas não recebidas</div><canvas id="inadPag"></canvas></div>
<div class="card col-6"><h2>Assinatura — inadimplência por método</h2><div class="desc">PIX Automático concentra o risco</div><canvas id="assin"></canvas></div>
<div class="card col-12"><h2>Devedores por quantidade de parcelas em atraso</h2>
<div class="desc">Barras = clientes · Linha = valor em atraso · quem tem 1-3 parcelas é o mais recuperável</div><canvas id="distParc" style="max-height:300px"></canvas></div>
<div class="card col-12"><h2>Inadimplência por lançamento (% sobre vendas PARCELADAS)</h2>
<div class="desc">Parceladas = onde a Maré tem risco (Recuperador, Assinatura, Parcelado Hotmart, Parcelamento Inteligente). Vendas à vista pro produtor excluídas — não têm risco. Lançamentos com &lt;50 vendas não aparecem.</div>
<canvas id="lanc" style="max-height:320px"></canvas></div>
<div class="card col-12"><h2>📊 Comparativo por % Perda de Faturamento por vendedor (Código SRC)</h2>
<div class="desc">"% Perda" = de cada R$ 100 já vencidos, quanto virou prejuízo. Inclui todas as vendas (à vista e parcelado).</div>
<table><thead><tr><th>Origem</th><th>Vendas</th><th>Recebido</th><th>Em atraso</th><th>A receber</th><th>% Perda</th><th>% Recup.</th></tr></thead><tbody id="tblVend"></tbody></table></div>
<div class="card col-7"><h2>Valor em atraso por tipo de plano</h2>
<table><thead><tr><th>Tipo de plano</th><th>Vendas</th><th>Devedores</th><th>% inad.</th><th>Valor em atraso</th></tr></thead><tbody id="tblTipo"></tbody></table></div>
<div class="card col-5"><h2>Top 12 maiores devedores</h2>
<table><thead><tr><th>Cliente</th><th>Parc.</th><th>Em atraso</th></tr></thead><tbody id="tblDev"></tbody></table></div>
</div><div class="foot">Gerado automaticamente do histórico Hotmart · Maré Educação · ref. __REF__</div></div>
<script>
const D=__DATA__;
Chart.defaults.color='#94a3b8';Chart.defaults.borderColor='#334155';
const BRL=v=>'R$ '+Number(v).toLocaleString('pt-BR',{maximumFractionDigits:0});
const k=D.kpi,mi=v=>(v/1e6).toLocaleString('pt-BR',{maximumFractionDigits:2}),mil=v=>Math.round(v/1000);
document.getElementById('kpis').innerHTML=`
<div class="kpi green"><div class="label">Já recebido (líquido)</div><div class="val">R$ ${mi(k.recebido)} mi</div><div class="sub">${k.pagas.toLocaleString('pt-BR')} parcelas pagas</div></div>
<div class="kpi"><div class="label">A receber — Otimista 12m</div><div class="val">R$ ${mi(k.otim)} mi</div><div class="sub">${k.qfut.toLocaleString('pt-BR')} parcelas futuras</div></div>
<div class="kpi violet"><div class="label">A receber — Ajustado 12m</div><div class="val">R$ ${mil(k.ajus)} mil</div><div class="sub">−inadimplência histórica</div></div>
<div class="kpi red"><div class="label">Em atraso confirmado</div><div class="val">R$ ${mil(k.atraso)} mil</div><div class="sub">${k.qatraso.toLocaleString('pt-BR')} parcelas não pagas</div></div>
<div class="kpi amber"><div class="label">Inadimplência geral</div><div class="val">${k.taxa}%</div><div class="sub">das parcelas vencidas</div></div>
<div class="kpi red"><div class="label">Clientes devedores</div><div class="val">${k.devedores}</div><div class="sub">de ${k.alunos.toLocaleString('pt-BR')} alunos (${k.pctdev}%)</div></div>`;
const projC=new Chart(proj,{type:'bar',data:{labels:D.mLbl,datasets:[
{label:'Otimista (100%)',data:D.otim,backgroundColor:'#38bdf8',borderRadius:5},
{label:'Ajustado (realista)',data:D.ajus,backgroundColor:'#a78bfa',borderRadius:5}]},
options:{plugins:{legend:{position:'top'},tooltip:{callbacks:{label:c=>c.dataset.label+': '+BRL(c.raw)}}},scales:{y:{ticks:{callback:v=>'R$ '+(v/1000)+'k'}}}}});
const acumC=new Chart(document.getElementById('acum'),{type:'line',data:{labels:D.mLbl,datasets:[{label:'Acumulado',data:D.ac.slice(),borderColor:'#22c55e',backgroundColor:'rgba(34,197,94,.15)',fill:true,tension:.35,pointRadius:3,spanGaps:false}]},
options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>BRL(c.raw)}}},scales:{y:{ticks:{callback:v=>'R$ '+(v/1000)+'k'}}}}});
new Chart(hist,{type:'bar',data:{labels:D.hLbl,datasets:[{label:'Recebido',data:D.hTot,backgroundColor:'#0ea5e9',borderRadius:5}]},
options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>BRL(c.raw)}}},scales:{y:{ticks:{callback:v=>'R$ '+(v/1000)+'k'}}}}});
const pal=['#ef4444','#f97316','#f59e0b','#eab308','#84cc16','#22c55e','#10b981','#06b6d4'];
new Chart(inadTipo,{type:'bar',data:{labels:D.itLbl,datasets:[{data:D.itVal,backgroundColor:D.itLbl.map((_,i)=>pal[i%pal.length]),borderRadius:5}]},
options:{indexAxis:'y',plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.raw+'%'}}},scales:{x:{ticks:{callback:v=>v+'%'}}}}});
new Chart(inadPag,{type:'bar',data:{labels:D.ipLbl,datasets:[{data:D.ipVal,backgroundColor:'#fb7185',borderRadius:5}]},
options:{indexAxis:'y',plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.raw+'%'}}},scales:{x:{ticks:{callback:v=>v+'%'}}}}});
new Chart(assin,{type:'bar',data:{labels:D.asLbl,datasets:[{label:'% inad.',data:D.asVal,backgroundColor:D.asLbl.map((_,i)=>pal[i%pal.length]),borderRadius:5}]},
options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.raw+'%'}}},scales:{y:{ticks:{callback:v=>v+'%'}}}}});
new Chart(distParc,{data:{labels:D.dLbl,datasets:[
{type:'bar',label:'Clientes',data:D.dCli,backgroundColor:D.dLbl.map((l,i)=>['#22c55e','#4ade80','#86efac','#fbbf24','#f59e0b','#f97316','#fb923c','#ef4444','#dc2626','#b91c1c','#7f1d1d'][Math.min(i,10)]),borderRadius:5,yAxisID:'y'},
{type:'line',label:'Valor em atraso',data:D.dVal,borderColor:'#38bdf8',backgroundColor:'rgba(56,189,248,.15)',fill:true,tension:.35,pointRadius:3,yAxisID:'y1'}]},
options:{plugins:{legend:{position:'top'},tooltip:{callbacks:{label:c=>c.dataset.type==='line'?'Valor: '+BRL(c.raw):'Clientes: '+c.raw}}},
scales:{x:{title:{display:true,text:'Qtd. de parcelas em atraso por cliente'}},y:{position:'left',title:{display:true,text:'Clientes'}},y1:{position:'right',grid:{drawOnChartArea:false},ticks:{callback:v=>'R$ '+(v/1000)+'k'}}}}});
// Lançamentos — barras Total/Parceladas/Inad + linha de % inad sobre parcelado
new Chart(lanc,{data:{labels:D.lLbl.map((m,i)=>D.lRot[i]?D.lRot[i]+' ('+m+')':m),datasets:[
{type:'bar',label:'Total de vendas',data:D.lTot,backgroundColor:'#475569',borderRadius:4,yAxisID:'y'},
{type:'bar',label:'Parceladas (risco)',data:D.lParc,backgroundColor:'#a78bfa',borderRadius:4,yAxisID:'y'},
{type:'bar',label:'Inadimplentes',data:D.lInad,backgroundColor:'#ef4444',borderRadius:4,yAxisID:'y'},
{type:'line',label:'% inad. sobre parceladas',data:D.lPct,borderColor:'#fbbf24',backgroundColor:'rgba(251,191,36,.15)',fill:false,tension:.3,pointRadius:5,pointBackgroundColor:'#fbbf24',yAxisID:'y1'}]},
options:{plugins:{legend:{position:'top'},tooltip:{callbacks:{label:c=>c.dataset.type==='line'?c.dataset.label+': '+c.raw+'%':c.dataset.label+': '+c.raw}}},
scales:{x:{title:{display:true,text:'Lançamentos'}},y:{position:'left',title:{display:true,text:'Vendas'},beginAtZero:true},y1:{position:'right',grid:{drawOnChartArea:false},title:{display:true,text:'% inad.'},ticks:{callback:v=>v+'%'},beginAtZero:true,max:100}}}});
const TG={'t-hi':v=>v>=25,'t-md':v=>v>=10};
D.qt.forEach(r=>{let p=parseFloat(r[3]),c=p>=25?'t-hi':p>=10?'t-md':'t-lo';
tblTipo.insertAdjacentHTML('beforeend',`<tr><td>${r[0]}</td><td class=num>${r[1]}</td><td class=num>${r[2]}</td><td><span class="tag ${c}">${r[3]}</span></td><td class=num>${BRL(r[4])}</td></tr>`)});
D.top.forEach(r=>tblDev.insertAdjacentHTML('beforeend',`<tr><td>${r[0]}</td><td class=num>${r[1]}</td><td class=num>R$ ${Number(r[2]).toLocaleString('pt-BR',{minimumFractionDigits:2})}</td></tr>`));
D.vend.forEach(r=>{const p=r[5];const c=p>=25?'#ef4444':p>=12?'#f59e0b':'#22c55e';const dot=p>=25?'🔴':p>=12?'🟡':'🟢';
tblVend.insertAdjacentHTML('beforeend',`<tr><td>${dot} ${r[0]}</td><td class=num>${r[1].toLocaleString('pt-BR')}</td><td class=num>${BRL(r[2])}</td><td class=num style="color:#fb7185">${BRL(r[3])}</td><td class=num style="color:#a78bfa">${BRL(r[4])}</td><td class=num style="color:${c};font-weight:700">${p}%</td><td class=num>${r[6]}%</td></tr>`)});
try{const fDe=document.getElementById('fDe'),fAte=document.getElementById('fAte');
D.mLbl.forEach((m,i)=>{fDe.add(new Option(m,i));fAte.add(new Option(m,i))});
fDe.value=0;fAte.value=D.mLbl.length-1;
const ap=()=>{let a=+fDe.value,b=+fAte.value;if(a>b){[a,b]=[b,a];fDe.value=a;fAte.value=b}
const dd=i=>i>=a&&i<=b;
projC.data.datasets[0].backgroundColor=D.mLbl.map((_,i)=>dd(i)?'#38bdf8':'rgba(56,189,248,.15)');
projC.data.datasets[1].backgroundColor=D.mLbl.map((_,i)=>dd(i)?'#a78bfa':'rgba(167,139,250,.15)');projC.update();
let s=0;acumC.data.datasets[0].data=D.ajus.map((v,i)=>dd(i)?(s+=v):null);acumC.update();
const sl=ar=>ar.reduce((x,v,i)=>x+(dd(i)?v:0),0);
rOt.textContent=BRL(sl(D.otim));rAj.textContent=BRL(sl(D.ajus));rQt.textContent=sl(D.qtd).toLocaleString('pt-BR')}
fDe.onchange=ap;fAte.onchange=ap;fReset.onclick=()=>{fDe.value=0;fAte.value=D.mLbl.length-1;ap()};ap();
}catch(e){console.error('Filtro:',e)}
</script></body></html>'''

_html = _DASH_TMPL.replace('__DATA__', _json.dumps(D, ensure_ascii=False)).replace('__REF__', D['ref'])
with open(OUT_DASH, 'w', encoding='utf-8') as _f:
    _f.write(_html)
print(f"  - Dashboard dinâmico: {OUT_DASH}")
